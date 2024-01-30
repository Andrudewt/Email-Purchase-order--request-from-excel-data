import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pretty_html_table import build_table
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time


def contacts(path):
    """" Sorts address book file """
    wb = load_workbook(path)
    ws = wb['contacts']
    column_1 = [sell.value for sell in ws['A']]
    column_2 = [sell.value for sell in ws['B']]
    address_book = pd.DataFrame({'address': column_2[1:]}, index=column_1[1:])
    return address_book['address']


def send_email(my_email, password, tbl: 'dataframe', mail: 'email') -> None:
    """Sends email"""
    server = smtplib.SMTP('smtp.gmail.com', 587)  # 587 - port
    server.starttls()
    time.sleep(1)
    server.login(my_email, password)  # Gain access

    msg = MIMEMultipart()  # Create message
    msg['From'] = my_email  # Sender

    msg['Subject'] = 'Request for quotation'
    msg['Return-Receipt-To'] = my_email

    msg['To'] = mail  # Recipient
    body = ("<!DOCTYPE html><head></head><body>Hello, dear " + mail +
            '. Could you please send your commercial offer for: ' + "</body></html>")  # Message body
    msg.attach(MIMEText(body, 'html'))
    msg.attach(MIMEText(tbl, 'html'))
    try:
        server.send_message(msg)
    except Exception:
        print("error ", mail)
    else:
        print("sent")

    time.sleep(40)  # Latency to avoid ban
    server.quit()


def sort_and_send(my_email, password, path: 'excel_file', address_book):
    wb = load_workbook(path)
    ws = wb.worksheets[0]

    data = ws.values
    columns = next(data)[0:]
    df = pd.DataFrame(data, columns=columns)
    # Get the list of proceeded purchase requests
    requests = df['Purchase request'].value_counts()

    # Data consolidation (equal by quantity)
    df = df.groupby(['art'], as_index=False).agg(
        {'Manufacturer': 'first', 'Title': 'first', 'quant': 'sum', 'st.un': 'first'})[
        ['Manufacturer', 'art', 'Title', 'quant', 'st.un']].sort_values('Manufacturer')

    # Items that don't have supplier assignment go to the relevant sheet
    wb.create_sheet('Not_assigned', 1)  # create new sheet
    ws1 = wb['Not_assigned']

    address_book = contacts(address_book)

    # Split the sorted data into tables by Supplier and send to relevant address
    for i in df['Manufacturer'].unique():
        data = (df[df['Manufacturer'] == str(i)].sort_values(['Title'], ascending=[True]))
        table = build_table(data, 'blue_light')
        try:
            send_email(my_email, password, table, address_book[i])
        except:
            for r in dataframe_to_rows(data, index=False, header=True):
                ws1.append(r)

    # Putting done purchase requests into next sheet
    wb.create_sheet('Performed requests', 2)
    ws2 = wb['Performed requests']
    requests = np.array(requests.index.tolist()).reshape(len(requests), 1)
    for r in requests:
        ws2.append(tuple(r))

    wb.save(path)


def main():
    excel_file_path = 'Input_data.xlsx'   # Data file path
    address_book = 'address_book.xlsx'    # Address book file path
    my_email = input('Enter your email address: ')     # Enter sender email information
    password = input('Enter your email password: ')
    sort_and_send(my_email, password, excel_file_path, address_book)


if __name__ == '__main__':
    main()

